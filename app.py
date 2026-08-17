import streamlit as st
import pandas as pd
import datetime
import openpyxl
import requests
import io
import plotly.graph_objects as go
from plotly.subplots import make_subplots

# Configuración de la página
st.set_page_config(page_title="Dashboard Planta - Balance Metalúrgico", layout="wide")

# Ocultar menú superior y footer de Streamlit
st.markdown("""
    <style>
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    </style>
""", unsafe_allow_html=True)

# -------------------------------------------------------------
# 1. DESCARGA EN VIVO DESDE GOOGLE SHEETS
# -------------------------------------------------------------
SHEET_ID = "16qHmnhtgGDETeOCeahGZ-Ka_8a2d0KYH"
URL_DESCARGA = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"

@st.cache_data(ttl=60)
def obtener_bytes_excel(url: str) -> io.BytesIO:
    response = requests.get(url, timeout=15)
    response.raise_for_status()
    return io.BytesIO(response.content)

try:
    excel_stream = obtener_bytes_excel(URL_DESCARGA)
except Exception as e:
    st.error(f"Error al conectar con Google Sheets: {e}")
    st.stop()

# Parámetros Económicos
CU_PRICE = 11084.00
PB_PRICE = 1981.00
ZN_PRICE = 3041.00
AG_PRICE = 53.05
FACTOR_TN_LBS = 2204.62

# -------------------------------------------------------------
# 2. LECTURA DE HOJAS (DASHBOARD 1: BALANCE METALÚRGICO)
# -------------------------------------------------------------
# A) MES PPT
excel_stream.seek(0)
df_ppt_raw = pd.read_excel(excel_stream, sheet_name='MES PPT')
df_ppt_raw.columns = df_ppt_raw.iloc[0]
df_ppt = df_ppt_raw.iloc[1:].reset_index(drop=True)

mensual_lbs_dict = {}
for _, row in df_ppt.iterrows():
    fecha_val = row.get('FECHA')
    if pd.notna(fecha_val):
        try:
            f_str = pd.to_datetime(fecha_val).strftime("%d/%m/%Y")
        except Exception:
            continue
            
        zn_ppt = float(row.get('Fines Zn_EZ', 0.0) or 0.0)
        pb_ppt = float(row.get('Fines Pb_EP', 0.0) or 0.0)
        cu_ppt = float(row.get('Fines Cu_EC', 0.0) or 0.0)
        ag_cu_ppt = float(row.get('Fines Ag_EC', 0.0) or 0.0)
        ag_pb_ppt = float(row.get('Fines Ag_EP', 0.0) or 0.0)
        ag_zn_ppt = float(row.get('Fines Ag_EZ', row.get('Oz/t Ag_EZ', 0.0)) or 0.0)

        val_zn = zn_ppt * ZN_PRICE
        val_pb = pb_ppt * PB_PRICE
        val_ag = (ag_cu_ppt + ag_pb_ppt + ag_zn_ppt) * AG_PRICE
        val_totales = val_zn + val_pb + val_ag
        cueq_sin_cu = val_totales / CU_PRICE
        cueq_tms = cueq_sin_cu + cu_ppt
        lbs_cueq_ppt = (cueq_tms * FACTOR_TN_LBS) / 1000.0

        mensual_lbs_dict[f_str] = lbs_cueq_ppt

# B) TRATAMIENTO
excel_stream.seek(0)
xl = pd.ExcelFile(excel_stream)
sheet_trat_name = next((s for s in xl.sheet_names if 'TRAT' in s.upper()), None)
mensual_trat_dict, ejec_trat_dict = {}, {}

if sheet_trat_name:
    df_trat = pd.read_excel(excel_stream, sheet_name=sheet_trat_name)
    for _, row in df_trat.iterrows():
        fecha_val = row.get('FECHA')
        if pd.notna(fecha_val):
            try:
                f_str = pd.to_datetime(fecha_val).strftime("%d/%m/%Y")
            except Exception:
                continue
            ejec_val = row.get('EJECTUADO Lb Cu Eq', row.get('EJECUTADO Lb Cu Eq', row.get('EJEC TRAT', 0.0)))
            mens_val = row.get('MENSUAL Lb Cu Eq', row.get('MENSUAL TRAT', 0.0))
            ejec_trat_dict[f_str] = float(ejec_val) if pd.notna(ejec_val) else 0.0
            mensual_trat_dict[f_str] = float(mens_val) if pd.notna(mens_val) else 0.0

# C) AGOSTO
excel_stream.seek(0)
wb_in = openpyxl.load_workbook(excel_stream, data_only=True)
sheet_agosto_name = next((s for s in wb_in.sheetnames if 'AGOSTO' in s.strip().upper()), 'AGOSTO')
ws_agosto = wb_in[sheet_agosto_name]

agosto_data_dict = {}
def clean_val(val):
    try: return float(val) if val is not None else 0.0
    except: return 0.0

for r in range(1, ws_agosto.max_row + 1):
    val_a = ws_agosto.cell(row=r, column=1).value
    if not val_a: continue
    f_str = None
    if isinstance(val_a, (datetime.datetime, datetime.date)):
        f_str = val_a.strftime("%d/%m/%Y")
    else:
        try: f_str = pd.to_datetime(str(val_a).strip(), dayfirst=True).strftime("%d/%m/%Y")
        except: continue
    if not f_str: continue

    zn_tms, pb_tms, cu_tms, ag_cu, ag_pb, ag_zn = 0.0, 0.0, 0.0, 0.0, 0.0, 0.0
    for sub_r in range(r, min(r + 15, ws_agosto.max_row + 1)):
        prod_raw = str(ws_agosto.cell(row=sub_r, column=1).value or '').replace(" ", "").upper()
        if 'CONC.CU' in prod_raw:
            cu_tms = clean_val(ws_agosto.cell(row=sub_r, column=9).value)
            ag_cu  = clean_val(ws_agosto.cell(row=sub_r, column=12).value)
        elif 'CONC.PB' in prod_raw:
            pb_tms = clean_val(ws_agosto.cell(row=sub_r, column=10).value)
            ag_pb  = clean_val(ws_agosto.cell(row=sub_r, column=12).value)
        elif 'CONC.ZN' in prod_raw:
            zn_tms = clean_val(ws_agosto.cell(row=sub_r, column=11).value)
            ag_zn  = clean_val(ws_agosto.cell(row=sub_r, column=12).value)

    agosto_data_dict[f_str] = {
        'cu_tms': cu_tms, 'pb_tms': pb_tms, 'zn_tms': zn_tms,
        'ag_cu': ag_cu, 'ag_pb': ag_pb, 'ag_zn': ag_zn
    }

# -------------------------------------------------------------
# 3. TRANSFORMACIÓN Y TABLA BALANCE METALÚRGICO
# -------------------------------------------------------------
base_date = datetime.date(2026, 7, 26)
date_list = [base_date + datetime.timedelta(days=i) for i in range(31)]
rows_data = []

for d in date_list:
    fecha_fmt = d.strftime("%d/%m/%Y")
    fecha_label = d.strftime("%d - %b")

    data_day = agosto_data_dict.get(fecha_fmt, {'cu_tms': 0.0, 'pb_tms': 0.0, 'zn_tms': 0.0, 'ag_cu': 0.0, 'ag_pb': 0.0, 'ag_zn': 0.0})
    cu_tms, pb_tms, zn_tms = data_day['cu_tms'], data_day['pb_tms'], data_day['zn_tms']
    ag_cu, ag_pb, ag_zn = data_day['ag_cu'], data_day['ag_pb'], data_day['ag_zn']

    val_zn = zn_tms * ZN_PRICE
    val_pb = pb_tms * PB_PRICE
    val_ag = (ag_cu + ag_pb + ag_zn) * AG_PRICE
    val_totales = val_zn + val_pb + val_ag
    cueq_sin_cu = val_totales / CU_PRICE
    cueq_tms = cueq_sin_cu + cu_tms
    lbs_cueq = cueq_tms * FACTOR_TN_LBS / 1000.0

    m_lbs = mensual_lbs_dict.get(fecha_fmt, 0.0)
    m_trat = mensual_trat_dict.get(fecha_fmt, 0.0)
    e_trat = ejec_trat_dict.get(fecha_fmt, 0.0)

    rows_data.append([
        fecha_fmt, fecha_label, zn_tms, pb_tms, cu_tms, ag_cu, ag_pb, ag_zn,
        val_zn, val_pb, val_ag, val_totales, cueq_sin_cu, cueq_tms,
        FACTOR_TN_LBS, lbs_cueq, m_lbs, m_trat, e_trat
    ])

headers = [
    'Fecha', 'Etiqueta', 'Zn (TMS)', 'Pb (TMS)', 'Cu (TMS)', 'Ag en Cu (Oz)', 'Ag en Pb (Oz)', 'Ag en Zn (Oz)',
    '$ Zn', '$ Pb', '$ Ag', '$ TOTALES', 'CuEq sin Cont. Metal. de Cu', 'Cu Equivalente TMS',
    'Factor converion Tn a Lbs', 'Lbs Cu Eq', 'MENSUAL Lb Cu Eq', 'MENSUAL TRAT', 'EJEC TRAT'
]
df_res = pd.DataFrame(rows_data, columns=headers)

# Cálculos KPI 1
df_ejecutado = df_res[df_res['Lbs Cu Eq'] > 0]
dias_con_datos = len(df_ejecutado)

ejecutado_total = df_ejecutado['Lbs Cu Eq'].sum()
mensual_total = df_res['MENSUAL Lb Cu Eq'].sum()

mensual_a_la_fecha = df_res.iloc[:dias_con_datos]['MENSUAL Lb Cu Eq'].sum()
proyeccion_restante = df_res.iloc[dias_con_datos:]['MENSUAL Lb Cu Eq'].sum()
ejec_mas_proy = ejecutado_total + proyeccion_restante

cumplimiento_pct = (ejecutado_total / mensual_a_la_fecha * 100) if mensual_a_la_fecha > 0 else 0
promedio_diario = (ejecutado_total / dias_con_datos) if dias_con_datos > 0 else 0

# -------------------------------------------------------------
# 4. RENDERIZAR DASHBOARD 1: BALANCE METALÚRGICO
# -------------------------------------------------------------
st.markdown("<h2 style='text-align: center; color: white;'>Reporte Diario de Planta - Balance Metalúrgico</h2>", unsafe_allow_html=True)

text_mensual_lbs = [f"{x:.0f}" if x > 0 else "" for x in df_res['MENSUAL Lb Cu Eq']]
text_ejec_lbs = [f"{x:.0f}" if x > 0 else "" for x in df_res['Lbs Cu Eq']]
text_mensual_trat = [f"{x:,.0f}" if x > 0 else "" for x in df_res['MENSUAL TRAT']]
text_ejec_trat = [f"{x:,.0f}" if x > 0 else "" for x in df_res['EJEC TRAT']]

fig_main = make_subplots(specs=[[{"secondary_y": True}]])

fig_main.add_trace(go.Bar(
    x=df_res['Etiqueta'], y=df_res['MENSUAL Lb Cu Eq'], name='MENSUAL Lb Cu Eq', marker_color='#2A629A',
    text=text_mensual_lbs, textposition='inside', textfont=dict(size=9, color='white')
), secondary_y=False)

fig_main.add_trace(go.Bar(
    x=df_res['Etiqueta'], y=df_res['Lbs Cu Eq'], name='EJECUTADO Lb Cu Eq', marker_color='#FF7F3E',
    text=text_ejec_lbs, textposition='inside', textfont=dict(size=9, color='white')
), secondary_y=False)

fig_main.add_trace(go.Scatter(
    x=df_res['Etiqueta'], y=df_res['MENSUAL TRAT'], name='PLAN MENSUAL TMS TRAT', mode='lines+markers+text',
    line=dict(color='#FF2E63', width=2), marker=dict(size=4),
    text=text_mensual_trat, textposition='top center', textfont=dict(size=8, color='#FF2E63')
), secondary_y=True)

ejec_trat_masked = [x if x > 0 else None for x in df_res['EJEC TRAT']]
fig_main.add_trace(go.Scatter(
    x=df_res['Etiqueta'], y=ejec_trat_masked, name='EJECUTADO TMS TRAT', mode='lines+markers+text',
    line=dict(color='#FFD700', width=2), marker=dict(size=5),
    text=text_ejec_trat, textposition='bottom center', textfont=dict(size=8, color='#FFD700')
), secondary_y=True)

fig_main.update_layout(
    template='plotly_dark',
    height=480,
    barmode='group',
    legend=dict(orientation='h', yanchor='bottom', y=1.05, xanchor='center', x=0.5, font=dict(size=11)),
    margin=dict(l=30, r=30, t=40, b=30),
    paper_bgcolor='rgba(0,0,0,0)',
    plot_bgcolor='rgba(0,0,0,0)'
)
fig_main.update_xaxes(tickangle=-45, tickfont=dict(size=10))
fig_main.update_yaxes(title_text="<b>Lb Cu Eq (Miles)</b>", secondary_y=False, showgrid=True, gridcolor='#333333')
fig_main.update_yaxes(title_text="<b>Tratamiento (TMS)</b>", secondary_y=True, showgrid=False)

st.plotly_chart(fig_main, use_container_width=True)

st.write("") 

col1, col2, col3, col4 = st.columns([1.1, 1.0, 0.9, 0.9])

with col1:
    fig_mensual = go.Figure()
    fig_mensual.add_trace(go.Bar(
        x=['EJEC + PROY', 'EJECUTADO', 'MENSUAL'], y=[ejec_mas_proy, ejecutado_total, mensual_total],
        marker_color=['#E63946', '#FFB703', '#1D3557'],
        text=[f"{ejec_mas_proy:.0f}", f"{ejecutado_total:.0f}", f"{mensual_total:.0f}"], textposition='outside',
        textfont=dict(size=11, color='white')
    ))
    fig_mensual.update_layout(
        title=dict(text="<b>CUMPLIMIENTO MENSUAL</b>", x=0.5, font=dict(size=12, color='white')),
        template='plotly_dark', height=260, margin=dict(l=10, r=10, t=35, b=10),
        paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
        yaxis=dict(showticklabels=False, showgrid=False, range=[0, max(ejec_mas_proy, mensual_total) * 1.25])
    )
    st.plotly_chart(fig_mensual, use_container_width=True)

with col2:
    fig_a_fecha = go.Figure()
    fig_a_fecha.add_trace(go.Bar(
        x=['MENSUAL', 'EJECUTADO'], y=[mensual_a_la_fecha, ejecutado_total],
        marker_color=['#1D3557', '#FFB703'],
        text=[f"{mensual_a_la_fecha:.0f}", f"{ejecutado_total:.0f}"], textposition='outside',
        textfont=dict(size=11, color='white')
    ))
    fig_a_fecha.update_layout(
        title=dict(text="<b>CUMPLIMIENTO A LA FECHA</b>", x=0.5, font=dict(size=12, color='white')),
        template='plotly_dark', height=260, margin=dict(l=10, r=10, t=35, b=10),
        paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
        yaxis=dict(showticklabels=False, showgrid=False, range=[0, max(mensual_a_la_fecha, ejecutado_total) * 1.25])
    )
    st.plotly_chart(fig_a_fecha, use_container_width=True)

with col3:
    st.markdown(f"""
        <div style="background-color: #1A2536; border: 1px solid #2E3E52; border-radius: 8px; padding: 20px 10px; text-align: center; margin-top: 20px;">
            <div style="font-size: 13px; font-weight: bold; color: #8A99AD; margin-bottom: 8px;">CUMPLIMIENTO A LA FECHA</div>
            <div style="font-size: 40px; font-weight: 900; color: #38BDF8; font-family: 'Arial Black';">{cumplimiento_pct:.0f}%</div>
        </div>
    """, unsafe_allow_html=True)

with col4:
    st.markdown(f"""
        <div style="background-color: #1A2536; border: 1px solid #2E3E52; border-radius: 8px; padding: 20px 10px; text-align: center; margin-top: 20px;">
            <div style="font-size: 13px; font-weight: bold; color: #8A99AD; margin-bottom: 8px;">PROMEDIO EJECUTADO DIARIO</div>
            <div style="font-size: 40px; font-weight: 900; color: #38BDF8; font-family: 'Arial Black';">{promedio_diario:.0f}</div>
        </div>
    """, unsafe_allow_html=True)


# =============================================================
# 5. LECTURA Y PROCESAMIENTO - MÓDULO DE PRODUCCIÓN (DASHBOARD 2)
# =============================================================
st.markdown("<br><hr style='border-color: #333333;'><br>", unsafe_allow_html=True)

# Lector de pestañas 'PLANEAMIENTO' y 'PRODUCCION'
try:
    excel_stream.seek(0)
    df_plan_raw = pd.read_excel(excel_stream, sheet_name='PLANEAMIENTO')
    excel_stream.seek(0)
    df_prod_raw = pd.read_excel(excel_stream, sheet_name='PRODUCCION')

    # Renombrar TMS-ACUMULADO a EJECUTADO
    if 'TMS-ACUMULADO' in df_prod_raw.columns:
        df_prod_raw = df_prod_raw.rename(columns={'TMS-ACUMULADO': 'EJECUTADO'})

    # Formatear fechas
    df_plan_raw['FECHA'] = pd.to_datetime(df_plan_raw['FECHA'])
    df_prod_raw['FECHA'] = pd.to_datetime(df_prod_raw['FECHA'])

    # Fusión por fecha
    df_prod_mix = pd.merge(df_plan_raw, df_prod_raw[['FECHA', 'EJECUTADO']], on='FECHA', how='left')

    # Actualización de EJEC + PROYEC: Si hay EJECUTADO > 0, se asigna EJECUTADO; de lo contrario se deja el proyectado
    df_prod_mix['EJEC + PROYEC'] = df_prod_mix.apply(
        lambda r: r['EJECUTADO'] if (pd.notna(r['EJECUTADO']) and r['EJECUTADO'] > 0) else r['EJEC + PROYEC'],
        axis=1
    )

    df_prod_mix['Etiqueta'] = df_prod_mix['FECHA'].dt.strftime('%d - %b')

    # Cálculos KPI Producción
    df_p_ejec = df_prod_mix[df_prod_mix['EJECUTADO'] > 0]
    dias_p_ejec = len(df_p_ejec)

    p_ejecutado_total = df_p_ejec['EJECUTADO'].sum()
    p_ejec_proy_total = df_prod_mix['EJEC + PROYEC'].sum()
    p_mensual_total = df_prod_mix['MENSUAL'].sum()

    p_mensual_a_la_fecha = df_prod_mix.iloc[:dias_p_ejec]['MENSUAL'].sum() if dias_p_ejec > 0 else 0
    p_cumplimiento_pct = (p_ejecutado_total / p_mensual_a_la_fecha * 100) if p_mensual_a_la_fecha > 0 else 0
    p_promedio_diario = (p_ejecutado_total / dias_p_ejec) if dias_p_ejec > 0 else 0

    # -------------------------------------------------------------
    # 6. RENDERIZAR DASHBOARD 2: PRODUCCIÓN (TMS)
    # -------------------------------------------------------------
    st.markdown("<h2 style='text-align: center; color: white;'>PRODUCCIÓN (TMS)</h2>", unsafe_allow_html=True)

    text_p_ejec = [f"{x:,.0f}" if (pd.notna(x) and x > 0) else "" for x in df_prod_mix['EJECUTADO']]
    text_p_sem = [f"{x:,.0f}" if (pd.notna(x) and x > 0) else "" for x in df_prod_mix['SEMANAL']]
    text_p_mens = [f"{x:,.0f}" if (pd.notna(x) and x > 0) else "" for x in df_prod_mix['MENSUAL']]

    fig_p_main = go.Figure()

    # Barras EJECUTADO
    fig_p_main.add_trace(go.Bar(
        x=df_prod_mix['Etiqueta'], y=df_prod_mix['EJECUTADO'], name='EJECUTADO', marker_color='#2A629A',
        text=text_p_ejec, textposition='inside', textfont=dict(size=9, color='white')
    ))

    # Línea SEMANAL
    fig_p_main.add_trace(go.Scatter(
        x=df_prod_mix['Etiqueta'], y=df_prod_mix['SEMANAL'], name='SEMANAL', mode='lines+markers+text',
        line=dict(color='#FFD700', width=2), marker=dict(size=5),
        text=text_p_sem, textposition='top center', textfont=dict(size=8, color='#FFD700')
    ))

    # Línea MENSUAL
    fig_p_main.add_trace(go.Scatter(
        x=df_prod_mix['Etiqueta'], y=df_prod_mix['MENSUAL'], name='MENSUAL', mode='lines+markers+text',
        line=dict(color='#FF2E63', width=2), marker=dict(size=4),
        text=text_p_mens, textposition='bottom center', textfont=dict(size=8, color='#FF2E63')
    ))

    fig_p_main.update_layout(
        template='plotly_dark',
        height=480,
        legend=dict(orientation='h', yanchor='bottom', y=1.05, xanchor='center', x=0.5, font=dict(size=11)),
        margin=dict(l=30, r=30, t=40, b=30),
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)'
    )
    fig_p_main.update_xaxes(tickangle=-45, tickfont=dict(size=10))
    fig_p_main.update_yaxes(title_text="<b>Producción (TMS)</b>", showgrid=True, gridcolor='#333333')

    st.plotly_chart(fig_p_main, use_container_width=True)

    st.write("") 

    # Cuadros inferiores Producción (TMS)
    p_col1, p_col2, p_col3, p_col4 = st.columns([1.1, 1.0, 0.9, 0.9])

    with p_col1:
        fig_p_mensual = go.Figure()
        fig_p_mensual.add_trace(go.Bar(
            x=['EJEC + PROY', 'EJECUTADO', 'MENSUAL'], y=[p_ejec_proy_total, p_ejecutado_total, p_mensual_total],
            marker_color=['#E63946', '#FFB703', '#1D3557'],
            text=[f"{p_ejec_proy_total:,.0f}", f"{p_ejecutado_total:,.0f}", f"{p_mensual_total:,.0f}"],
            textposition='outside', textfont=dict(size=11, color='white')
        ))
        fig_p_mensual.update_layout(
            title=dict(text="<b>PRODUCCIÓN (TMS)</b>", x=0.5, font=dict(size=12, color='white')),
            template='plotly_dark', height=260, margin=dict(l=10, r=10, t=35, b=10),
            paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
            yaxis=dict(showticklabels=False, showgrid=False, range=[0, max(p_ejec_proy_total, p_mensual_total, 1.0) * 1.25])
        )
        st.plotly_chart(fig_p_mensual, use_container_width=True)

    with p_col2:
        fig_p_a_fecha = go.Figure()
        fig_p_a_fecha.add_trace(go.Bar(
            x=['MENSUAL', 'EJECUTADO'], y=[p_mensual_a_la_fecha, p_ejecutado_total],
            marker_color=['#1D3557', '#FFB703'],
            text=[f"{p_mensual_a_la_fecha:,.0f}", f"{p_ejecutado_total:,.0f}"],
            textposition='outside', textfont=dict(size=11, color='white')
        ))
        fig_p_a_fecha.update_layout(
            title=dict(text="<b>CUMPLIMIENTO A LA FECHA</b>", x=0.5, font=dict(size=12, color='white')),
            template='plotly_dark', height=260, margin=dict(l=10, r=10, t=35, b=10),
            paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
            yaxis=dict(showticklabels=False, showgrid=False, range=[0, max(p_mensual_a_la_fecha, p_ejecutado_total, 1.0) * 1.25])
        )
        st.plotly_chart(fig_p_a_fecha, use_container_width=True)

    with p_col3:
        st.markdown(f"""
            <div style="background-color: #1A2536; border: 1px solid #2E3E52; border-radius: 8px; padding: 20px 10px; text-align: center; margin-top: 20px;">
                <div style="font-size: 13px; font-weight: bold; color: #8A99AD; margin-bottom: 8px;">CUMPLIMIENTO A LA FECHA</div>
                <div style="font-size: 40px; font-weight: 900; color: #38BDF8; font-family: 'Arial Black';">{p_cumplimiento_pct:.0f}%</div>
            </div>
        """, unsafe_allow_html=True)

    with p_col4:
        st.markdown(f"""
            <div style="background-color: #1A2536; border: 1px solid #2E3E52; border-radius: 8px; padding: 20px 10px; text-align: center; margin-top: 20px;">
                <div style="font-size: 13px; font-weight: bold; color: #8A99AD; margin-bottom: 8px;">PROMEDIO EJECUTADO DIARIO</div>
                <div style="font-size: 40px; font-weight: 900; color: #38BDF8; font-family: 'Arial Black';">{p_promedio_diario:,.0f}</div>
            </div>
        """, unsafe_allow_html=True)

except Exception as e:
    st.warning(f"No se pudo cargar la sección de Producción/Planeamiento: {e}")
